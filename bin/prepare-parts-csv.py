#!/usr/bin/env python3
"""
Convert Lista_Piese.xlsx + PIESE/ images into a WooCommerce-compatible CSV.

Usage:
    python3 bin/prepare-parts-csv.py

Input:
    ~/Downloads/Lista_Piese.xlsx
    ~/Downloads/PIESE/

Output:
    ~/Downloads/woocommerce-parts-import.csv
    ~/Downloads/parts-images/          (flat dir, COD-prefixed JPGs)
    ~/Downloads/unmatched-report.txt   (products without matched images)
"""

import csv
import os
import re
import subprocess
import sys
import unicodedata
from difflib import SequenceMatcher
from pathlib import Path

# Auto-install dependencies
def ensure_deps():
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        print("Installing openpyxl...")
        subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'openpyxl'])

    try:
        from PIL import Image  # noqa: F401
    except ImportError:
        print("Installing Pillow...")
        subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'Pillow'])

ensure_deps()

import openpyxl
from PIL import Image

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

EXCEL_PATH = Path.home() / "Downloads" / "Lista_Piese.xlsx"
IMAGES_DIR = Path.home() / "Downloads" / "PIESE"
OUTPUT_CSV = Path.home() / "Downloads" / "woocommerce-parts-import.csv"
OUTPUT_IMAGES = Path.home() / "Downloads" / "parts-images"
OUTPUT_REPORT = Path.home() / "Downloads" / "unmatched-report.txt"

# Base URL where images will be uploaded on the server
IMAGE_BASE_URL = "https://dev.trotibike.ro/wp-content/uploads/parts-images"

# Placeholder image URL for products without matched images
PLACEHOLDER_IMAGE = ""

# Map Excel sheet names to model display names
SHEET_TO_MODEL = {
    'G2':           'Kukirin G2',
    'G2 MASTER':    'Kukirin G2 Master',
    'G3 PRO':       'Kukirin G3 Pro',
    'G4':           'Kukirin G4',
    'G4 MAX':       'Kukirin G4 Max',
    'S1 MAX':       'Kukirin S1 Max',
    'A1':           'Kukirin A1',
    'M4 MAX':       'Kukirin M4 Max',
    'C1 PRO':       'Kukirin C1 Pro',
    'G2 PRO 2023':  'Kukirin G2 Pro 2023',
    'G2 PRO 2024':  'Kukirin G2 Pro 2024',
    'G2 MAX':       'Kukirin G2 Max',
    'G3':           'Kukirin G3',
}

# Map Excel sheet names to PIESE/ subfolder names (case-sensitive filesystem match)
SHEET_TO_IMAGE_FOLDER = {
    'G2':           'G2',
    'G2 MASTER':    'G2 Master',
    'G3 PRO':       'G3 Pro',
    'G4':           'G4',
    'G4 MAX':       'G4 Max',
    'S1 MAX':       'S1 Max',
    'A1':           'A1',
    'M4 MAX':       'M4 max',
    'C1 PRO':       'C1 Pro',
    'G2 PRO 2023':  'G2 Pro',
    'G2 PRO 2024':  'G2 Pro',
    'G2 MAX':       'G2 max',
    'G3':           'G3',
}

# Model short prefixes to strip from product names (longest first for matching)
MODEL_PREFIXES = [
    'EU24-G2pro', 'EUG2pro-G2max', 'EUG2pro',
    'G2 Master', 'G2 MASTER', 'G2master', 'G2Master',
    'G3 Pro', 'G3 PRO', 'G3pro', 'G3Pro',
    'G4 Max', 'G4 MAX', 'G4max', 'G4Max',
    'S1 Max', 'S1 MAX', 'S1max', 'S1Max',
    'M4 Max', 'M4 MAX', 'M4max', 'M4Max',
    'C1 Pro', 'C1 PRO', 'C1pro', 'C1Pro',
    'G2 Pro', 'G2 PRO',
    'G2 Max', 'G2 MAX', 'G2max', 'G2Max',
    'G2', 'G3', 'G4', 'A1',
]

# Allowed source image extensions
IMAGE_EXTENSIONS = {'.jpg', '.jpeg', '.png', '.webp', '.avif'}

# Words to strip for fuzzy matching
STOP_WORDS = {
    'pentru', 'kukirin', 'kugoo', 'kirin', 'trotineta', 'electrica',
    'original', 'originala', 'originale', 'model',
    'de', 'din', 'cu', 'si', 'la', 'pe', 'in', 'a', 'al', 'ale',
    'g2', 'g3', 'g4', 'a1', 's1', 'm4', 'c1',
    'g2max', 'g2master', 'g3pro', 'g4max', 's1max', 'm4max', 'c1pro',
    'kukirin', 'kugoo',
    'max', 'pro', 'master',
}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def strip_diacritics(text: str) -> str:
    """Remove Romanian diacritics and normalize unicode."""
    nfkd = unicodedata.normalize('NFKD', text)
    return ''.join(c for c in nfkd if not unicodedata.combining(c))


def normalize_for_match(name: str) -> str:
    """Normalize a name for fuzzy matching: strip diacritics, lowercase, remove stop words."""
    s = strip_diacritics(name.strip().lower())
    # Remove content in parentheses like (A), (B), (C), (pereche)
    s = re.sub(r'\([^)]*\)', '', s)
    # Split on non-alpha
    words = re.split(r'[^a-z0-9]+', s)
    words = [w for w in words if w and w not in STOP_WORDS and len(w) > 1]
    return ' '.join(words)


def strip_model_prefix(name: str) -> str:
    """Strip model prefix from a product name (e.g., 'G2 Crossbar' -> 'Crossbar')."""
    stripped = name.strip()
    for prefix in sorted(MODEL_PREFIXES, key=len, reverse=True):
        if stripped.startswith(prefix):
            stripped = stripped[len(prefix):].strip(' -–—/')
            break
        if stripped.lower().startswith(prefix.lower()):
            stripped = stripped[len(prefix):].strip(' -–—/')
            break
    return stripped


def fuzzy_score(product_name: str, folder_name: str) -> float:
    """Score how well a product name matches a folder name (0.0 - 1.0)."""
    norm_p = normalize_for_match(strip_model_prefix(product_name))
    norm_f = normalize_for_match(folder_name)

    if not norm_p or not norm_f:
        return 0.0

    # Exact match after normalization
    if norm_p == norm_f:
        return 1.0

    # Containment check
    if norm_p in norm_f or norm_f in norm_p:
        return 0.85

    # Word overlap score
    words_p = set(norm_p.split())
    words_f = set(norm_f.split())
    if words_p and words_f:
        overlap = len(words_p & words_f)
        total = max(len(words_p), len(words_f))
        word_score = overlap / total if total else 0
    else:
        word_score = 0

    # Sequence similarity
    seq_score = SequenceMatcher(None, norm_p, norm_f).ratio()

    # Combined score (word overlap weighted higher)
    return max(word_score * 0.7 + seq_score * 0.3, seq_score)


def find_best_image_folder(product_name: str, folder_names: list, cutoff: float = 0.35) -> tuple:
    """Find the best matching image folder. Returns (folder_name, score) or (None, 0)."""
    best_folder = None
    best_score = 0.0

    for folder in folder_names:
        score = fuzzy_score(product_name, folder)
        if score > best_score:
            best_score = score
            best_folder = folder

    if best_score >= cutoff:
        return best_folder, best_score
    return None, 0.0


def convert_to_jpg(src_path: Path, dst_path: Path) -> bool:
    """Convert any image to JPG format with quality optimization."""
    try:
        with Image.open(src_path) as img:
            # Convert to RGB (JPEG doesn't support alpha/palette)
            if img.mode in ('RGBA', 'LA', 'P', 'PA'):
                background = Image.new('RGB', img.size, (255, 255, 255))
                if img.mode == 'P':
                    img = img.convert('RGBA')
                if 'A' in img.mode:
                    background.paste(img, mask=img.split()[-1])
                else:
                    background.paste(img)
                img = background
            elif img.mode != 'RGB':
                img = img.convert('RGB')

            # Resize if very large (keep reasonable for web)
            max_size = 1200
            if img.width > max_size or img.height > max_size:
                img.thumbnail((max_size, max_size), Image.LANCZOS)

            img.save(dst_path, 'JPEG', quality=85, optimize=True)
            return True
    except Exception as e:
        print(f"  WARNING: Failed to convert {src_path.name}: {e}")
        return False


def get_image_files(folder_path: Path) -> list:
    """Get all image files from a folder (non-recursive), sorted by name."""
    if not folder_path.is_dir():
        return []

    images = []
    for f in folder_path.iterdir():
        if f.is_file() and f.suffix.lower() in IMAGE_EXTENSIONS:
            # Skip very small files (likely corrupt)
            if f.stat().st_size > 1000:
                images.append(f)

    images.sort(key=lambda f: f.name.lower())
    return images


def clean_product_name(ro_name: str, model_name: str) -> str:
    """Clean Romanian product name: strip model prefix, append full model name."""
    name = strip_model_prefix(ro_name)
    # Clean up leading/trailing punctuation
    name = name.strip(' -–—/')
    # Capitalize first letter
    if name:
        name = name[0].upper() + name[1:]
    return f"{name} {model_name}" if name else model_name


def parse_sheet(ws, sheet_name: str) -> list:
    """Parse a single Excel sheet and return list of product dicts."""
    products = []
    model_name = SHEET_TO_MODEL.get(sheet_name)
    if not model_name:
        return products

    # Read header row (row 3) to detect column layout
    header_row = []
    for row in ws.iter_rows(min_row=3, max_row=3, values_only=True):
        header_row = [str(v).strip().upper() if v else '' for v in row]
        break

    # G2 sheet has 2x "DENUMIRE ENGLEZA" columns (duplicate EN name)
    # All other sheets have: EN | RO | COD | PRICE_PARTNER | PRICE_RETAIL [| empty]
    has_double_en = (
        len(header_row) >= 2
        and 'ENGLEZA' in header_row[0]
        and 'ENGLEZA' in header_row[1]
    )

    for row in ws.iter_rows(min_row=4, values_only=True):
        vals = list(row)

        if has_double_en:
            # 6-col: EN1 | EN2 | RO | COD | PRICE_PARTNER | PRICE_RETAIL
            en_name = str(vals[1]).strip() if vals[1] else (str(vals[0]).strip() if vals[0] else '')
            ro_name = str(vals[2]).strip() if len(vals) > 2 and vals[2] else ''
            cod = vals[3] if len(vals) > 3 else None
            price_retail = vals[5] if len(vals) > 5 else None
        else:
            # 5-col: EN | RO | COD | PRICE_PARTNER | PRICE_RETAIL
            en_name = str(vals[0]).strip() if vals[0] else ''
            ro_name = str(vals[1]).strip() if len(vals) > 1 and vals[1] else ''
            cod = vals[2] if len(vals) > 2 else None
            price_retail = vals[4] if len(vals) > 4 else None

        # Skip non-data rows
        if not en_name or not cod:
            continue

        # Validate COD is a number
        try:
            cod_num = int(float(cod))
        except (ValueError, TypeError):
            continue

        if cod_num < 20000 or cod_num > 50000:
            continue

        if not ro_name:
            ro_name = en_name  # Fallback to English name

        # Parse price
        price = None
        if price_retail:
            try:
                price = int(float(price_retail))
                if price <= 0:
                    price = None
            except (ValueError, TypeError):
                pass

        products.append({
            'cod': str(cod_num),
            'name_en': en_name,
            'name_ro': ro_name,
            'price': price,
            'model': model_name,
            'sheet': sheet_name,
        })

    return products


def build_folder_index(images_dir: Path) -> dict:
    """
    Build an index of all image folders: {model_folder: {subfolder_name: Path}}.
    Uses the subfolder names (product descriptions) as keys for matching.
    """
    index = {}
    if not images_dir.is_dir():
        return index

    for model_dir in images_dir.iterdir():
        if not model_dir.is_dir():
            continue
        model_name = model_dir.name
        index[model_name] = {}
        for subfolder in model_dir.iterdir():
            if subfolder.is_dir():
                index[model_name][subfolder.name] = subfolder

    return index


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    if not EXCEL_PATH.exists():
        print(f"ERROR: Excel file not found: {EXCEL_PATH}")
        sys.exit(1)

    has_images = IMAGES_DIR.exists()
    if not has_images:
        print(f"WARNING: Images directory not found: {IMAGES_DIR}")
        print("Products will be created without images.\n")

    # Clean and create output directory
    if OUTPUT_IMAGES.exists():
        import shutil
        shutil.rmtree(OUTPUT_IMAGES)
    OUTPUT_IMAGES.mkdir(parents=True, exist_ok=True)

    # Parse Excel
    print(f"Reading {EXCEL_PATH}...")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

    all_products = []
    for sheet_name in wb.sheetnames:
        if sheet_name not in SHEET_TO_MODEL:
            continue
        ws = wb[sheet_name]
        products = parse_sheet(ws, sheet_name)
        all_products.extend(products)
        print(f"  {sheet_name:20s} {len(products):3d} products")

    print(f"\nTotal products parsed: {len(all_products)}")

    # Check for duplicate CODs
    cods = [p['cod'] for p in all_products]
    seen_cods = set()
    dupes = set()
    for c in cods:
        if c in seen_cods:
            dupes.add(c)
        seen_cods.add(c)

    if dupes:
        print(f"WARNING: {len(dupes)} duplicate CODs found.")
        print(f"  Duplicates will have model suffix appended to SKU.")
        # Deduplicate: append model short code to SKU for dupes
        cod_count = {}
        for p in all_products:
            cod_count[p['cod']] = cod_count.get(p['cod'], 0) + 1
        for p in all_products:
            if p['cod'] in dupes:
                # Append a short model suffix: "25112" -> "25112-G2M" for G2 Master
                short = p['sheet'].replace(' ', '').replace('PRO', 'P').replace('MAX', 'X')[:4]
                p['sku'] = f"{p['cod']}-{short}"
            else:
                p['sku'] = p['cod']

    else:
        for p in all_products:
            p['sku'] = p['cod']

    # Build image folder index
    folder_index = build_folder_index(IMAGES_DIR) if has_images else {}

    # Match images
    print(f"\nMatching images...")
    stats = {'matched': 0, 'unmatched': 0, 'images_converted': 0}
    unmatched = []

    for product in all_products:
        sheet_name = product['sheet']
        img_folder_key = SHEET_TO_IMAGE_FOLDER.get(sheet_name, '')

        # Get the subfolders for this model
        model_subfolders = folder_index.get(img_folder_key, {})

        if not model_subfolders:
            product['images'] = []
            product['image_match'] = 'no_folder'
            stats['unmatched'] += 1
            unmatched.append(product)
            continue

        # Try to match product name to a subfolder
        subfolder_names = list(model_subfolders.keys())
        matched_name, score = find_best_image_folder(product['name_ro'], subfolder_names)

        if matched_name:
            folder_path = model_subfolders[matched_name]
            image_files = get_image_files(folder_path)

            if image_files:
                product['images'] = []
                for idx, img_file in enumerate(image_files, 1):
                    out_name = f"{product['sku']}-{idx}.jpg"
                    out_path = OUTPUT_IMAGES / out_name

                    if convert_to_jpg(img_file, out_path):
                        product['images'].append(out_name)
                        stats['images_converted'] += 1

                product['image_match'] = f'matched({score:.2f})'
                product['image_folder'] = matched_name
                stats['matched'] += 1
            else:
                product['images'] = []
                product['image_match'] = 'empty_folder'
                stats['unmatched'] += 1
                unmatched.append(product)
        else:
            product['images'] = []
            product['image_match'] = 'no_match'
            stats['unmatched'] += 1
            unmatched.append(product)

    print(f"\n  Matched:          {stats['matched']}")
    print(f"  Unmatched:        {stats['unmatched']}")
    print(f"  Images converted: {stats['images_converted']}")

    # Write unmatched report
    with open(OUTPUT_REPORT, 'w', encoding='utf-8') as f:
        f.write("Unmatched Products Report\n")
        f.write("========================\n\n")
        f.write(f"Total unmatched: {len(unmatched)} / {len(all_products)}\n")
        f.write(f"Total matched:   {stats['matched']} / {len(all_products)}\n\n")

        by_model = {}
        for p in unmatched:
            by_model.setdefault(p['model'], []).append(p)

        for model, prods in sorted(by_model.items()):
            f.write(f"\n--- {model} ({len(prods)} unmatched) ---\n")
            # Also list available folders for this model
            sheet = prods[0]['sheet']
            img_key = SHEET_TO_IMAGE_FOLDER.get(sheet, '')
            available = sorted(folder_index.get(img_key, {}).keys())
            if available:
                f.write(f"    Available folders: {len(available)}\n")
            f.write("\n")
            for p in prods:
                norm = normalize_for_match(strip_model_prefix(p['name_ro']))
                f.write(f"  COD {p['cod']}: {p['name_ro']}\n")
                f.write(f"    normalized: {norm}\n")
                f.write(f"    reason: {p['image_match']}\n\n")

    print(f"\nUnmatched report: {OUTPUT_REPORT}")

    # Generate CSV
    print(f"\nGenerating WooCommerce CSV...")

    csv_rows = []
    for product in all_products:
        name = clean_product_name(product['name_ro'], product['model'])
        price = product['price'] if product['price'] else ''
        published = 1 if product['price'] and product['price'] > 0 else 0

        # Image URLs — use matched images or placeholder
        if product['images']:
            image_urls = ','.join(
                f"{IMAGE_BASE_URL}/{img}" for img in product['images']
            )
        else:
            image_urls = PLACEHOLDER_IMAGE

        csv_rows.append({
            'Name': name,
            'Type': 'simple',
            'SKU': product['sku'],
            'Published': published,
            'Regular price': price,
            'Categories': f"Piese de schimb > {product['model']}",
            'Tags': product['model'],
            'Short description': product['name_en'],
            'Description': '',
            'Tax status': 'taxable',
            'In stock?': 1,
            'Stock': '',
            'Images': image_urls,
            'Meta: _spare_part_model': product['model'],
            'Meta: _spare_part_source': 'kukirin_excel',
        })

    fieldnames = [
        'Name', 'Type', 'SKU', 'Published', 'Regular price',
        'Categories', 'Tags', 'Short description', 'Description',
        'Tax status', 'In stock?', 'Stock', 'Images',
        'Meta: _spare_part_model', 'Meta: _spare_part_source',
    ]

    with open(OUTPUT_CSV, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(csv_rows)

    # Summary
    with_price = sum(1 for p in all_products if p['price'] and p['price'] > 0)
    with_images = sum(1 for p in all_products if p['images'])

    print(f"\n{'='*50}")
    print(f"SUMMARY")
    print(f"{'='*50}")
    print(f"Total products:    {len(all_products)}")
    print(f"With price:        {with_price} (will be published)")
    print(f"Without price:     {len(all_products) - with_price} (will be draft)")
    print(f"With images:       {with_images}")
    print(f"Without images:    {len(all_products) - with_images} (placeholder)")
    print(f"Images converted:  {stats['images_converted']} JPGs")
    print(f"\nOutput files:")
    print(f"  CSV:      {OUTPUT_CSV}")
    print(f"  Images:   {OUTPUT_IMAGES}/ ({stats['images_converted']} files)")
    print(f"  Report:   {OUTPUT_REPORT}")
    print(f"\nNext steps:")
    print(f"  1. Review unmatched report and manually add images if needed")
    print(f"  2. Upload {OUTPUT_IMAGES}/ contents to Hostinger:")
    print(f"     wp-content/uploads/parts-images/")
    print(f"  3. Import CSV via WooCommerce > Products > Import")


if __name__ == '__main__':
    main()
